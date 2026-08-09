"""
Excel/CSV 文件解析器 —— v1 数据分析链路的入口。

职责：
1. 读取 .xlsx / .xls / .csv 文件
2. 推断每列的数据类型（用于 NL2SQL 生成 CREATE TABLE 语句）
3. 处理合并单元格（前向填充 None 值）
4. 多 sheet 支持
5. 返回结构化的 ParseResult

设计要点：
- 列类型推断采用"采样 + 逐行尝试"策略，不确定时回退到 string
- 合并单元格：openpyxl 只在左上角保留值，其余为 None → 做前向填充
- CSV 用 pandas 读取（自动处理编码、分隔符），Excel 用 openpyxl
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.schemas import ColumnMeta, ColumnType, ParseResult, SheetInfo


# 采样行数上限：大文件只采样前 N 行做类型推断，避免全量扫描
MAX_SAMPLE_ROWS = 1000

# 布尔值识别：列中所有非空唯一值落在这个集合里 → 判定为 boolean
BOOL_VALUES: set[str | bool | int] = {
    True, False, 0, 1,
    "true", "false", "yes", "no", "是", "否",
    "True", "False", "Yes", "No", "TRUE", "FALSE", "YES", "NO",
}


class ExcelParser:
    """Excel/CSV 解析器 —— 无状态，纯函数式设计。"""

    @staticmethod
    def parse(file_path: str | Path) -> ParseResult:
        """入口：根据扩展名分派到 Excel 或 CSV 解析路径。"""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        suffix = file_path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return ExcelParser._parse_excel(file_path)
        elif suffix == ".csv":
            return ExcelParser._parse_csv(file_path)
        else:
            raise ValueError(f"不支持的文件类型: {suffix}")

    # ================================================================
    # Excel 解析
    # ================================================================

    @staticmethod
    def _parse_excel(file_path: Path) -> ParseResult:
        wb = load_workbook(file_path, data_only=True)
        sheets: list[SheetInfo] = []

        for ws in wb.worksheets:
            if ws.max_row < 2:  # 至少要有表头 + 一行数据
                continue
            sheet_info = ExcelParser._parse_worksheet(ws)
            sheets.append(sheet_info)

        wb.close()
        return ParseResult(
            file_name=file_path.name,
            file_type=file_path.suffix.lstrip("."),
            sheets=sheets,
        )

    @staticmethod
    def _parse_worksheet(ws: Worksheet) -> SheetInfo:
        # 第一行是表头
        headers: list[str] = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            headers.append(str(val).strip() if val is not None else f"col_{col_idx}")

        # 读取数据行
        raw_rows: list[dict[str, Any]] = []
        for row_idx in range(2, ws.max_row + 1):
            row: dict[str, Any] = {}
            for col_idx, header in enumerate(headers, start=1):
                row[header] = ws.cell(row=row_idx, column=col_idx).value
            raw_rows.append(row)

        # 合并单元格填充：只在实际合并区域内做填充，不误伤真正的缺失值
        rows = ExcelParser._fill_merged_cells(raw_rows, headers, ws)

        # 列类型推断
        columns = ExcelParser._infer_column_types(rows, headers)

        return SheetInfo(
            name=ws.title,
            columns=columns,
            records=rows,
        )

    # ================================================================
    # CSV 解析
    # ================================================================

    @staticmethod
    def _parse_csv(file_path: Path) -> ParseResult:
        # pandas 自动处理编码检测和分隔符推断
        df = pd.read_csv(file_path, nrows=None)

        headers = list(df.columns)
        # 统一列名为字符串、去空格
        headers = [str(h).strip() for h in headers]

        # NaN → None（和 Excel 路径统一用 None 表示缺失值）
        df = df.where(pd.notna(df), None)

        # 采样（大文件只取前 MAX_SAMPLE_ROWS 做类型推断）
        sample_df = df.head(MAX_SAMPLE_ROWS)
        rows: list[dict[str, Any]] = sample_df.to_dict(orient="records")

        columns = ExcelParser._infer_column_types(rows, headers)

        return ParseResult(
            file_name=file_path.name,
            file_type="csv",
            sheets=[
                SheetInfo(
                    name="data",  # CSV 只有一张"虚拟表"
                    columns=columns,
                    records=rows,
                )
            ],
        )

    # ================================================================
    # 合并单元格处理
    # ================================================================

    @staticmethod
    def _fill_merged_cells(
        rows: list[dict[str, Any]], headers: list[str], ws: Worksheet
    ) -> list[dict[str, Any]]:
        """只在 openpyxl 报告的合并区域内做前向填充。
        避免把真正的缺失值（None）误填为上一行的值。"""
        if not rows:
            return rows

        # 收集所有需要填充的单元格：(0-based row index, header) → 左上角的值
        fills: dict[tuple[int, str], Any] = {}
        for merged_range in ws.merged_cells.ranges:
            top_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    if r == merged_range.min_row and c == merged_range.min_col:
                        continue  # 跳过左上角（已有值）
                    data_row_idx = r - 2        # 转为 0-based rows 索引
                    col_header = headers[c - 1]  # 转为列名
                    if 0 <= data_row_idx < len(rows):
                        fills[(data_row_idx, col_header)] = top_val

        # 应用填充
        for (row_idx, header), val in fills.items():
            rows[row_idx][header] = val

        return rows

    # ================================================================
    # 列类型推断
    # ================================================================

    @staticmethod
    def _infer_column_types(
        rows: list[dict[str, Any]], headers: list[str]
    ) -> list[ColumnMeta]:
        """对每一列做类型推断，采样前 MAX_SAMPLE_ROWS 行。
        推断顺序：int → float → datetime → bool → string（越严格越优先）。"""
        sample_rows = rows[:MAX_SAMPLE_ROWS]
        columns: list[ColumnMeta] = []

        for header in headers:
            values = [row.get(header) for row in sample_rows]
            # 过滤 None 和空字符串
            non_null = [v for v in values if v is not None and v != ""]

            if not non_null:
                # 全空列 → 默认 string
                columns.append(ColumnMeta(
                    name=header, dtype=ColumnType.STRING,
                    nullable=True, unique_count=0, samples=[],
                ))
                continue

            dtype, unique_count = ExcelParser._guess_type(non_null)
            # 取最多 5 个样本值
            samples = list(dict.fromkeys(non_null))[:5]

            columns.append(ColumnMeta(
                name=header,
                dtype=dtype,
                nullable=len(non_null) < len(values),
                unique_count=unique_count,
                samples=samples,
            ))

        return columns

    @staticmethod
    def _guess_type(values: list[Any]) -> tuple[ColumnType, int]:
        """给定一列的非空值，返回 (最可能的类型, 唯一值数量)。"""
        unique = list(dict.fromkeys(values))
        unique_count = len(unique)

        # 1. 尝试 integer：全部能转成 int（排除带小数点的数字字符串）
        if all(ExcelParser._is_integer(v) for v in unique):
            return ColumnType.INTEGER, unique_count

        # 2. 尝试 float：全部能转成 float
        if all(ExcelParser._is_float(v) for v in unique):
            return ColumnType.FLOAT, unique_count

        # 3. 尝试 datetime：半数以上能解析为日期
        if ExcelParser._is_datetime_column(unique):
            return ColumnType.DATETIME, unique_count

        # 4. 尝试 boolean：所有非空值都在布尔候选集里
        if all(ExcelParser._is_bool(v) for v in unique):
            return ColumnType.BOOLEAN, unique_count

        # 5. 兜底 → string
        return ColumnType.STRING, unique_count

    # ================================================================
    # 类型判断辅助方法
    # ================================================================

    @staticmethod
    def _is_integer(value: Any) -> bool:
        """判断单个值是否可以安全转为 int。"""
        if isinstance(value, bool):
            return False  # bool 是 int 的子类，要先排除
        if isinstance(value, int):
            return True
        if isinstance(value, float):
            return value == int(value)  # 3.0 → True, 3.14 → False
        if isinstance(value, str):
            # 纯数字字符串（不含小数点），允许负号
            return bool(re.match(r"^-?\d+$", value.strip()))
        return False

    @staticmethod
    def _is_float(value: Any) -> bool:
        """判断单个值是否可以安全转为 float。"""
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            try:
                float(value.strip())
                return True
            except ValueError:
                return False
        return False

    @staticmethod
    def _is_datetime_column(values: list[Any]) -> bool:
        """判断一列是否主要为日期类型。
        策略：用 pandas.to_datetime 尝试解析，成功率 ≥ 60% 则认为是日期列。"""
        if len(values) < 2:
            return False
        # 纯数字列不判为日期（年份如 2024 可能被误判为日期）
        if all(isinstance(v, (int, float)) for v in values):
            return False
        str_values = [str(v).strip() for v in values]
        try:
            parsed = pd.to_datetime(str_values, errors="coerce")
            valid_ratio = parsed.notna().sum() / len(str_values)
            return valid_ratio >= 0.6
        except Exception:
            return False

    @staticmethod
    def _is_bool(value: Any) -> bool:
        """判断单个值是否属于布尔候选集。"""
        if isinstance(value, bool):
            return True
        if isinstance(value, str):
            return value.strip() in BOOL_VALUES
        if isinstance(value, int) and value in (0, 1):
            return True
        return False
