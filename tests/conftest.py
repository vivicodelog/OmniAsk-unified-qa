"""
pytest fixtures —— 自动生成测试数据文件，不依赖外部二进制文件。
"""

import tempfile
import time
from pathlib import Path
from typing import Iterator

import pymupdf
import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _safe_unlink(path: str) -> None:
    """安全删除临时文件——Windows 上 openpyxl 可能还没释放文件句柄。"""
    for _ in range(3):
        try:
            Path(path).unlink()
            return
        except PermissionError:
            time.sleep(0.1)
    # 最后尝试一次，失败了也无妨（临时目录会被系统清理）
    try:
        Path(path).unlink(missing_ok=True)
    except PermissionError:
        pass


@pytest.fixture
def sample_xlsx_path() -> Iterator[str]:
    """生成一个多 sheet 的 Excel 文件，覆盖所有列类型 + 合并单元格。

    Sheet "销售订单" 列：
    - 订单ID:    int       (1001, 1002, ...)
    - 客户名称:   string    (张三, 李四, ...)
    - 订单日期:   datetime  (2024-01-15, ...)
    - 金额:      float     (1500.50, ...)
    - 数量:      int       (2, 5, ...)
    - 是否会员:   bool      (是/否)
    - 备注:      string    (含 None 值)

    Sheet "客户信息" 列：
    - 客户ID:    int
    - 客户名称:   string
    - 注册日期:   datetime
    - 信用额度:   float (含 None)

    Sheet "合并区域" —— 测试合并单元格填充（使用真实的 merge_cells）
    """
    wb = Workbook()

    # ================================================================
    # Sheet 1: 销售订单
    # ================================================================
    ws1 = wb.active
    ws1.title = "销售订单"

    headers = ["订单ID", "客户名称", "订单日期", "金额", "数量", "是否会员", "备注"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    data = [
        [1001, "张三", "2024-01-15", 1500.50, 2, "是", "加急"],
        [1002, "李四", "2024-01-16", 3200.00, 5, "否", None],
        [1003, "王五", "2024-02-01", 899.99, 1, "是", "需要发票"],
        [1004, "赵六", "2024-02-10", 2100.00, 3, "否", None],
        [1005, "张三", "2024-03-01", 450.00, 10, "是", None],
        [1006, "钱七", "2024-03-15", 6780.50, 2, "是", "VIP客户"],
        [1007, "孙八", "2024-04-01", 1200.00, 4, "否", None],
        [1008, "李四", "2024-04-20", 3400.00, 1, "是", "退货处理中"],
        [1009, "周九", "2024-05-01", 560.00, 8, "否", None],
        [1010, "吴十", "2024-05-15", 9999.99, 1, "是", None],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=val)

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 16

    # ================================================================
    # Sheet 2: 客户信息
    # ================================================================
    ws2 = wb.create_sheet("客户信息")

    headers2 = ["客户ID", "客户名称", "注册日期", "信用额度"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)

    data2 = [
        [1, "张三", "2023-06-01", 50000.00],
        [2, "李四", "2023-08-15", 30000.00],
        [3, "王五", "2024-01-10", None],     # 信用额度缺失
        [4, "赵六", "2023-12-20", 20000.00],
        [5, "钱七", "2024-02-28", 100000.00],
    ]
    for row_idx, row_data in enumerate(data2, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=val)

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # ================================================================
    # Sheet 3: 合并区域 —— 使用真正的 merge_cells
    # ================================================================
    ws3 = wb.create_sheet("合并区域")

    headers3 = ["部门", "员工", "工资"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)

    # 先写数据
    merge_data = [
        ["技术部", "张三", 15000],
        ["技术部", "李四", 18000],
        ["技术部", "王五", 20000],
        ["销售部", "赵六", 12000],
        ["销售部", "钱七", 13000],
        ["市场部", "孙八", 14000],
    ]
    for row_idx, row_data in enumerate(merge_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws3.cell(row=row_idx, column=col_idx, value=val)

    # 再合并部门列——openpyxl 读回时只有左上角保留值
    ws3.merge_cells("A2:A4")   # 技术部 3 行
    ws3.merge_cells("A5:A6")   # 销售部 2 行
    # 市场部单独一行，不合并

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 16

    # ================================================================
    # 保存到临时文件
    # ================================================================
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()  # 关闭句柄，让 openpyxl 能写入
    wb.save(tmp.name)
    wb.close()
    yield tmp.name

    # 清理
    _safe_unlink(tmp.name)


@pytest.fixture
def sample_csv_path() -> Iterator[str]:
    """生成一个 CSV 测试文件。"""
    content = (
        "产品,销量,单价,日期,上架\n"
        "手机,100,2999.99,2024-01-15,是\n"
        "电脑,50,5999.00,2024-02-01,是\n"
        "平板,80,1999.50,2024-03-10,否\n"
        "耳机,200,299.00,2024-04-20,是\n"
        "手表,150,899.99,,否\n"  # 日期缺失
    )
    tmp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w", encoding="utf-8")
    tmp.write(content)
    tmp.close()
    yield tmp.name
    _safe_unlink(tmp.name)


@pytest.fixture
def sample_pdf_path() -> Iterator[str]:
    """生成一个含文本 + 图片的 PDF 测试文件（1 页）。"""
    doc = pymupdf.open()                # 新建空文档
    page = doc.new_page()               # 加一页（A4）

    # 插文本：中文必须 fontname="china-s"（内置 CJK 字体），默认 Helvetica 不支持中文
    page.insert_text((72, 72), "Hello 测试文本", fontname="china-s")

    # 插图片：Pixmap 现场造一张 10x10 纯灰图，不依赖外部图片文件
    pix = pymupdf.Pixmap(pymupdf.csRGB, pymupdf.IRect(0, 0, 10, 10), False)
    pix.clear_with(200)
    page.insert_image(pymupdf.Rect(100, 100, 200, 200), pixmap=pix)

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    tmp.close()
    doc.save(tmp.name)
    doc.close()
    yield tmp.name
    _safe_unlink(tmp.name)
